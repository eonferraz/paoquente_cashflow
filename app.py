import streamlit as st    
import pandas as pd
import pyodbc
import plotly.express as px
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime
import numpy as np
from plotly import graph_objects as go
import networkx as nx
import scipy
import calendar 
import plotly.io as pio
import datetime as dt
from openpyxl.styles import Alignment
import plotly.graph_objects as go

# CONFIG INICIAL
st.set_page_config(
    page_title="Pão Quente",
    layout="wide",
    initial_sidebar_state="collapsed"
)

#====================================================================================================================================
# CONEXÃO COM BANCO
@st.cache_data(ttl=300)
def carregar_dados():
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=sx-global.database.windows.net;'
        'DATABASE=sx_comercial;'
        'UID=paulo.ferraz;'
        'PWD=Gs!^42j$G0f0^EI#ZjRv'
    )
    df_vendas = pd.read_sql("SELECT * FROM PQ_VENDAS", conn)
    df_metas = pd.read_sql("SELECT * FROM PQ_METAS", conn)
    conn.close()
    return df_vendas, df_metas

#====================================================================================================================================
# SIDEBAR
def montar_sidebar(df, todas_uns):
    st.sidebar.markdown("## ⚙️ Painel de Controles")

    # Botão de recarregar
    if st.sidebar.button("🔄 Recarregar Dados", help="Atualiza os dados diretamente do banco"):
        st.cache_data.clear()
        st.experimental_rerun()

    st.sidebar.markdown("---")

    # Filtro de período
    data_ini = st.sidebar.date_input("📆 Data Início", value=df["DATA"].min())
    data_fim = st.sidebar.date_input("📆 Data Fim", value=df["DATA"].max())

    # Filtro de unidade
    un_selecionadas = st.sidebar.multiselect("🏬 Unidades:", todas_uns, default=todas_uns)

    st.sidebar.markdown("---")
    
    # Limpar filtros
    if st.sidebar.button("🧹 Limpar Filtros"):
        st.session_state.clear()
        st.experimental_rerun()

    return data_ini, data_fim, un_selecionadas

#====================================================================================================================================



#====================================================================================================================================
# CARGA E PREPARO
with st.spinner("🔄 Carregando dados..."):
    df, metas = carregar_dados()

# Limpeza e padronização
df.columns = df.columns.str.strip().str.upper()
metas.columns = metas.columns.str.strip().str.upper()

df["DATA"] = pd.to_datetime(df["DATA"], dayfirst=True, errors="coerce")
df = df.dropna(subset=["DATA"])
df["ANO_MES"] = df["DATA"].dt.to_period("M").astype(str)

metas["ANO_MES"] = pd.to_datetime(metas["ANO-MES"]).dt.to_period("M").astype(str)

# Agora sim, define todas_uns
todas_uns = sorted(metas["LOJA"].dropna().unique())

# Sidebar com controles
data_ini, data_fim, un_selecionadas = montar_sidebar(df, todas_uns)

# Aplica o filtro de data no DataFrame principal
df = df[(df["DATA"] >= pd.to_datetime(data_ini)) & (df["DATA"] <= pd.to_datetime(data_fim))]



# ====================
# CSS para o cabeçalho
# ====================
st.markdown("""
    <style>
        .fixed-header {
            position: sticky;
            top: 0;
            background-color: white;
            z-index: 999;
            padding: 10px 20px 5px 20px;
            border-bottom: 1px solid #ccc;
            box-shadow: 0px 2px 6px rgba(0,0,0,0.05);
        }
        .header-flex {
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .logo {
            height: 50px;
        }
        .title {
            font-size: 26px;
            font-weight: bold;
            color: #862E3A;
            margin: 0 auto;
        }
        .filters {
            display: flex;
            gap: 15px;
        }
        .block-container {
            padding-top: 0rem;
        }
    </style>
""", unsafe_allow_html=True)

# ====================
# HEADER ÚNICO COM LOGO E TÍTULO (sem filtros aqui)
# ====================
with st.container():
    st.markdown("<div class='fixed-header'>", unsafe_allow_html=True)
    st.markdown("<div class='header-flex'>", unsafe_allow_html=True)

    # Logo
    st.image("logo.png", width=90)

    # Título com logo embutido
    st.markdown("""
        <div style="display: flex; align-items: center; justify-content: center;">            
            <span style="font-size: 26px; font-weight: bold; color: #862E3A;">
                Dashboard Comercial - Pão Quente
            </span>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# ====================
# APLICAÇÃO DOS FILTROS (vindos do sidebar)
# ====================

# Filtro de datas (já aplicado antes ao df)
df_filt = df[df["UN"].isin(un_selecionadas)].copy()
metas_filt = metas[metas["LOJA"].isin(un_selecionadas)].copy()

# Geração automática dos meses presentes no df filtrado por data
meses_selecionados = sorted(df_filt["ANO_MES"].unique())
metas_filt = metas_filt[metas_filt["ANO_MES"].isin(meses_selecionados)]
df_filt = df_filt[df_filt["ANO_MES"].isin(meses_selecionados)]


# ====================
# DASHBOARD PRINCIPAL
# ====================

col1, col2 = st.columns([1, 4.2])

# CARDS
#=====================================================================================================================================================================
with col1:
    fat_total = df_filt["TOTAL"].sum()
    qtd_vendas = df_filt["COD_VENDA"].nunique()
    ticket = fat_total / qtd_vendas if qtd_vendas > 0 else 0
    meta_total = metas_filt["VALOR_META"].sum()
    progresso = (fat_total / meta_total) * 100 if meta_total > 0 else 0

    def metric_card(titulo, valor):
        st.markdown(
            f"""
            <div style="border: 1px solid #DDD; border-radius: 11px; padding: 11px; margin-bottom: 11px; text-align: center;">
                <div style="font-size: 13px; color: gray;">{titulo}</div>
                <div style="font-size: 22px; font-weight: bold;">{valor}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    # Formatação dos números com ponto para milhar e vírgula para decimal
    def format_brl(value):
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def format_percent(value):
        return f"{value:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")

    # Cards
    metric_card("💰 Faturamento Total", format_brl(fat_total))
    metric_card("🎯 Meta de Faturamento", format_brl(meta_total))
    metric_card("📈 Progresso da Meta", format_percent(progresso))
    metric_card("📊 Qtde de Vendas", f"{qtd_vendas:,}".replace(",", "."))
    metric_card("💳 Ticket Médio", format_brl(ticket))
#=====================================================================================================================================================================



# FATURAMENTO MENSAL
#=====================================================================================================================================================================
with col2:
    with st.container(border=True):
        # Preparação dos dados
        df_mes = df_filt.groupby("ANO_MES")["TOTAL"].sum().reset_index()
        df_meta_mes = metas_filt.groupby("ANO-MES")["VALOR_META"].sum().reset_index()
        df_meta_mes.rename(columns={"ANO-MES": "ANO_MES"}, inplace=True)

        # Merge mantendo todos os meses da meta (mesmo se não houve venda)
        df_merged = pd.merge(df_meta_mes, df_mes, on="ANO_MES", how="left").fillna(0)
        df_merged["PCT"] = df_merged["TOTAL"] / df_merged["VALOR_META"]

        # Gráfico de barras
        fig1 = px.bar(
            df_merged,
            x="ANO_MES",
            y=["VALOR_META", "TOTAL"],
            title="📊 Faturamento x Meta + % Realizado por Mês",
            barmode="group",
            color_discrete_sequence=["#A4B494", "#FE9C37"]
        )

        fig1.update_traces(
            texttemplate="R$ %{y:,.0f}",
            textposition="inside",
            textangle=-90,
            textfont_size=14,
            insidetextanchor="start"
        )

        # Linha de % realizado
        fig1.add_scatter(
            x=df_merged["ANO_MES"],
            y=df_merged["PCT"],
            mode="lines+markers",
            name="% Realizado",
            line=dict(color="#862E3A", dash="dot"),
            yaxis="y2",
            marker=dict(size=8)
        )

        # Linha fixa de referência em 100%
        fig1.add_shape(
            type="line",
            x0=df_merged["ANO_MES"].min(),
            x1=df_merged["ANO_MES"].max(),
            y0=1, y1=1,
            xref='x',
            yref='y2',
            line=dict(color="#C0392B", width=1.5, dash="dot")
        )

        # Destaque de % com cores vivas e acima da linha
        for i, row in df_merged.iterrows():
            cor_fundo = "#3CB371" if row["PCT"] >= 1 else "#C0392B"
            fig1.add_annotation(
                x=row["ANO_MES"],
                y=row["PCT"],
                text=f"{row['PCT']:.0%}",
                showarrow=False,
                font=dict(color="white", size=12),
                bgcolor=cor_fundo,
                borderpad=4,
                yanchor="top",
                yshift=-10
            )

        # Layout final com fundo transparente
        fig1.update_layout(
            template=pio.templates["plotly"],
            yaxis=dict(
                title="R$",
                tickprefix="R$ ",
                tickformat=",.0f",
                showticklabels=True,
                showgrid=False
            ),
            yaxis2=dict(
                overlaying="y",
                side="right",
                tickformat=".0%",
                title="%",
                range=[0, 1.5],
                showticklabels=True,
                showgrid=False
            ),
            xaxis=dict(
                type='category',
                tickangle=-45
            ),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="center",
                x=0.5
            ),
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)"
        )

        st.plotly_chart(fig1, use_container_width=True)
#=====================================================================================================================================================================


#st.markdown("---")


col3a, col3c, col3b = st.columns(3)  # Desempacota as colunas corretamente


#=====================================================================================================================================================================
# === BLOCOS DE GRÁFICOS COM CARDS INTERMEDIÁRIOS ===
from datetime import datetime
import calendar

# Cálculos
hoje = datetime.today()
dia_hoje = hoje.day
dias_no_mes = calendar.monthrange(hoje.year, hoje.month)[1]

df_merge = metas_filt.copy()
df_merge = df_merge.rename(columns={"LOJA": "UN", "VALOR_META": "VALOR_META"})

df_fat = df_filt.groupby("UN")["TOTAL"].sum().reset_index()
df_merge = pd.merge(df_merge, df_fat, on="UN", how="left").fillna(0)

# Faturamento acumulado e projetado
df_merge["FALTA_META"] = df_merge["VALOR_META"] - df_merge["TOTAL"]
df_merge["FALTA_META"] = df_merge["FALTA_META"].apply(lambda x: max(0, x))  # evita valores negativos
df_merge["MEDIA_DIARIA"] = df_merge["TOTAL"] / dia_hoje
df_merge["FAT_PROJETADO"] = df_merge["MEDIA_DIARIA"] * dias_no_mes
df_merge["PCT_PROJETADO"] = df_merge["FAT_PROJETADO"] / df_merge["VALOR_META"]

# ====================
# COLUNAS
# ====================
col3a, col3c, col_cards, col3b = st.columns([1.2, 1.2, 1.1, 1.2])

from datetime import datetime
import calendar

# === FILTRO PARA MÊS ATUAL
hoje = datetime.today()
ano_mes_atual = hoje.strftime("%Y-%m")
dia_hoje = hoje.day
dias_no_mes = calendar.monthrange(hoje.year, hoje.month)[1]

df_mes_atual = df_filt[df_filt["ANO_MES"] == ano_mes_atual].copy()
metas_mes_atual = metas_filt[metas_filt["ANO-MES"] == ano_mes_atual].copy()

# === PREPARAÇÃO DE DADOS AGRUPADOS POR UN
df_un_fat = df_mes_atual.groupby("UN")["TOTAL"].sum().reset_index()
df_merge = pd.merge(
    metas_mes_atual.rename(columns={"LOJA": "UN"}),
    df_un_fat,
    on="UN",
    how="left"
).fillna(0)

df_merge["FALTA_META"] = (df_merge["VALOR_META"] - df_merge["TOTAL"]).clip(lower=0)
df_merge["MEDIA_DIARIA"] = df_merge["TOTAL"] / dia_hoje
df_merge["FAT_PROJETADO"] = df_merge["MEDIA_DIARIA"] * dias_no_mes
df_merge["PCT_PROJETADO"] = df_merge["FAT_PROJETADO"] / df_merge["VALOR_META"]

# === CÁLCULOS PARA CARDS DO MÊS ATUAL
fat_realizado = df_merge["TOTAL"].sum()
meta = df_merge["VALOR_META"].sum()
fat_proj = df_merge["FAT_PROJETADO"].sum()
pct_proj = fat_proj / meta if meta > 0 else 0

# === LAYOUT EM 4 COLUNAS
col1, col2, col_card, col3 = st.columns([1.2, 1.2, 1.1, 1.2])

# === GRÁFICO 1 - Faturamento Atual vs Meta (Stacked)
with col1:
    with st.container(border=True):
        fig_fat = px.bar(
            df_merge,
            y="UN",
            x=["TOTAL", "FALTA_META"],
            orientation='h',
            barmode="stack",
            title="📊 Faturamento Atual vs Meta por UN",
            color_discrete_sequence=["#FE9C37", "#A4B494"],
            text_auto=".2s"
        )
        fig_fat.update_layout(
            xaxis_tickprefix="R$ ",
            xaxis_tickformat=",.0f",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5)
        )
        st.plotly_chart(fig_fat, use_container_width=True)

# === GRÁFICO 2 - Faturamento Projetado vs Meta
with col2:
    with st.container(border=True):
        fig_proj = px.bar(
            df_merge,
            x="UN",
            y=["VALOR_META", "FAT_PROJETADO"],
            barmode="group",
            title="🔮 Faturamento Projetado vs Meta",
            color_discrete_sequence=["#A4B494", "#37392E"],
            text_auto=".2s"
        )

        # Adiciona % como anotação
        for _, row in df_merge.iterrows():
            fig_proj.add_annotation(
                x=row["UN"],
                y=row["FAT_PROJETADO"],
                text=f"{row['PCT_PROJETADO']:.0%}",
                showarrow=False,
                yshift=10,
                font=dict(size=12, color="green" if row["PCT_PROJETADO"] >= 1 else "red")
            )

        fig_proj.update_layout(
            yaxis_tickprefix="R$ ",
            yaxis_tickformat=",.0f",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5)
        )
        st.plotly_chart(fig_proj, use_container_width=True)

# === CARDS CENTRAIS (apenas mês atual)
with col_card:
    with st.container(border=True):
        st.markdown("<h5 style='color:#862E3A; text-align:center;'>📋 Indicadores Gerais</h5>", unsafe_allow_html=True)
        st.metric("📈 Faturamento Realizado", f"R$ {fat_realizado:,.0f}".replace(",", "."))
        st.metric("🎯 Meta de Faturamento", f"R$ {meta:,.0f}".replace(",", "."))
        st.metric("🔮 Faturamento Projetado", f"R$ {fat_proj:,.0f}".replace(",", "."))
        st.metric("📊 Projeção vs Meta", f"{pct_proj:.0%}", delta="Acima da meta" if pct_proj >= 1 else "Abaixo da meta")

# === GRÁFICO 3 - Distribuição % por UN (não precisa de filtro por mês)
with col3:
    with st.container(border=True):
        fig_pie = px.pie(
            df_merge,
            names="UN",
            values="TOTAL",
            title="🍩 Distribuição % por UN",
            hole=0.5,
            color_discrete_sequence=px.colors.sequential.RdBu
        )
        fig_pie.update_traces(textposition="inside", textinfo="percent+label")
        st.plotly_chart(fig_pie, use_container_width=True)


#st.markdown("---")




# ANÁLISE DE PRODUTOS
with st.container(border=True):
    st.markdown("<h4 style='color:#862E3A;'>🏆 Top 10 Produtos e Produtos Associados</h4>", unsafe_allow_html=True)

    col1, col2 = st.columns([1.2, 1.8])

    # ================= COLUNA 1 - BARRAS =================
    with col1:
        df_top = df_filt.groupby("DESCRICAO_PRODUTO")["TOTAL"].sum().reset_index()
        df_top = df_top.sort_values("TOTAL", ascending=False).head(10)
        top_produtos = df_top["DESCRICAO_PRODUTO"].tolist()

        produto_selecionado = st.selectbox("🧠 Selecione um produto:", top_produtos)

        fig_top10 = px.bar(df_top.sort_values("TOTAL"),
                           x="TOTAL", y="DESCRICAO_PRODUTO",
                           orientation='h',
                           text_auto=True,
                           title="Top 10 Produtos",
                           color="TOTAL", color_continuous_scale="OrRd")

        fig_top10.update_layout(yaxis=dict(categoryorder="total ascending"),
                                xaxis_tickprefix="R$ ", xaxis_tickformat=",.2f",
                                margin=dict(t=40, l=10, r=10, b=10),
                                title_font=dict(size=16),
                                height=400)

        st.plotly_chart(fig_top10, use_container_width=True)

    # ================= COLUNA 2 - GRAFO =================
    with col2:
        df_assoc = df_filt[["COD_VENDA", "DESCRICAO_PRODUTO"]].drop_duplicates()

        vendas_com_produto = df_assoc[df_assoc["DESCRICAO_PRODUTO"] == produto_selecionado]["COD_VENDA"].unique()
        df_relacionados = df_assoc[df_assoc["COD_VENDA"].isin(vendas_com_produto)]

        total_vendas_produto = len(vendas_com_produto)

        relacionados = df_relacionados[df_relacionados["DESCRICAO_PRODUTO"] != produto_selecionado]
        freq_relacionados = relacionados["DESCRICAO_PRODUTO"].value_counts().head(5).reset_index()
        freq_relacionados.columns = ["PRODUTO", "FREQ"]
        freq_relacionados["PCT"] = freq_relacionados["FREQ"] / total_vendas_produto

        import networkx as nx
        import plotly.graph_objects as go

        G = nx.Graph()
        G.add_node(produto_selecionado, size=100)

        for _, row in freq_relacionados.iterrows():
            G.add_node(row["PRODUTO"], size=row["PCT"] * 100)
            G.add_edge(produto_selecionado, row["PRODUTO"], weight=row["PCT"])

        pos = nx.spring_layout(G, seed=42, k=0.8)

        edge_x, edge_y = [], []
        for edge in G.edges():
            x0, y0 = pos[edge[0]]
            x1, y1 = pos[edge[1]]
            edge_x += [x0, x1, None]
            edge_y += [y0, y1, None]

        edge_trace = go.Scatter(
            x=edge_x, y=edge_y,
            line=dict(width=1.5, color="#A4B494"),
            hoverinfo="none",
            mode="lines"
        )

        node_x, node_y, node_text, node_size = [], [], [], []
        for node in G.nodes():
            x, y = pos[node]
            node_x.append(x)
            node_y.append(y)

            if node == produto_selecionado:
                legenda = ""
                tamanho = 50
                texto = f"<b>{node}</b>"
            else:
                pct = G[produto_selecionado][node]['weight']
                legenda = f"em {pct:.1%} das vendas de {produto_selecionado}"
                tamanho = 20 + pct * 100
                texto = f"<b>{node}</b><br><span style='font-size:18px; color:#333;'>{legenda}</span>"
            
            node_text.append(texto)


            
            #if node == produto_selecionado:
            #    legenda = "Produto Selecionado"
            #    tamanho = 50
            #else:
            #    pct = G[produto_selecionado][node]['weight']
            #    legenda = f"{pct:.1%} das vendas com {produto_selecionado}"
            #    tamanho = 20 + pct * 100

            #node_text.append(f"<b>{node}</b><br><span style='font-size:13px; color:#333;'>{legenda}</span>")
            node_size.append(tamanho)

        node_trace = go.Scatter(
            x=node_x, y=node_y,
            mode="markers+text",
            hoverinfo="skip",
            text=node_text,
            textposition="bottom center",
            marker=dict(
                showscale=False,
                color=node_size,
                size=node_size,
                colorscale="OrRd",
                line_width=2
            )
        )

        fig_grafo = go.Figure(data=[edge_trace, node_trace],
                              layout=go.Layout(
                                  title=dict(text=f"Produtos Relacionados a: {produto_selecionado}", font=dict(size=16)),
                                  showlegend=False,
                                  margin=dict(t=40, l=0, r=0, b=0),
                                  hovermode="closest",
                                  xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                  yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                                  height=400
                              ))

        st.plotly_chart(fig_grafo, use_container_width=True)


# Análise por hora.
#===========================================================================================================================================================
with st.container(border=True):
    st.markdown("<h4 style='color:#862E3A;'>⏰ Desempenho de Vendas por Hora (com Drill-down por Período)</h4>", unsafe_allow_html=True)

    # Filtro de datas com início e fim
    min_data = df_filt["DATA"].min().date()
    max_data = df_filt["DATA"].max().date()
    hoje = datetime.today().date()

    col1, col2 = st.columns(2)
    with col1:
        data_inicio = st.date_input("📅 Data Início", value=hoje, min_value=min_data, max_value=max_data)
    with col2:
        data_fim = st.date_input("📅 Data Fim", value=hoje, min_value=min_data, max_value=max_data)

    df_periodo = df_filt[(df_filt["DATA"].dt.date >= data_inicio) & (df_filt["DATA"].dt.date <= data_fim)].copy()

    # Agrupamento por hora
    df_hora = df_periodo.groupby("HORA").agg({
        "TOTAL": "sum",
        "COD_VENDA": "nunique"
    }).reset_index().sort_values("HORA")

    df_hora["TICKET_MEDIO"] = df_hora["TOTAL"] / df_hora["COD_VENDA"]
    df_hora["HORA_STR"] = df_hora["HORA"].astype(str) + "h"
    media_total = df_hora["TOTAL"].mean()

    # Rótulo de dados: Faturamento + Ticket Médio
    df_hora["LABEL"] = df_hora.apply(
        lambda row: f"R$ {row['TOTAL']:,.2f}\n🎫 R$ {row['TICKET_MEDIO']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        axis=1
    )

    # Gráfico
    fig = go.Figure()

    # Barras - Faturamento
    fig.add_trace(go.Bar(
        x=df_hora["HORA_STR"],
        y=df_hora["TOTAL"],
        name="Faturamento",
        marker_color="#FE9C37",
        text=df_hora["LABEL"],
        textposition="outside",
        hovertemplate="<b>Hora:</b> %{x}<br><b>Faturamento:</b> R$ %{y:,.2f}<br><b>🎫 Ticket Médio:</b> %{text}<extra></extra>"
    ))

    # Linha - Quantidade de Vendas
    fig.add_trace(go.Scatter(
        x=df_hora["HORA_STR"],
        y=df_hora["COD_VENDA"],
        name="Qtd. Vendas",
        mode="lines+markers",
        marker=dict(color="#862E3A"),
        yaxis="y2",
        hovertemplate="Vendas: %{y}<extra></extra>"
    ))

    # Linha - Média de Faturamento
    fig.add_trace(go.Scatter(
        x=df_hora["HORA_STR"],
        y=[media_total] * len(df_hora),
        name="Média de Faturamento",
        mode="lines",
        line=dict(color="gray", dash="dot"),
        hoverinfo="skip"
    ))

    fig.update_layout(
        title="Desempenho por Hora",
        xaxis=dict(title="Hora do Dia", showgrid=False),
        yaxis=dict(
            title="Faturamento (R$)",
            tickprefix="R$ ",
            tickformat=",.0f",
            titlefont=dict(color="#FE9C37"),
            showgrid=False
        ),
        yaxis2=dict(
            title="Qtd. Vendas",
            overlaying="y",
            side="right",
            titlefont=dict(color="#862E3A"),
            showgrid=False
        ),
        legend=dict(orientation="h", y=1.02, x=0.5, xanchor="center", yanchor="bottom"),
        height=460,
        margin=dict(t=60, l=50, r=50, b=40)
    )

    st.plotly_chart(fig, use_container_width=True)

    # Exportar para Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas por Hora"
    ws.append(["Hora", "Faturamento", "Qtd. Vendas", "Ticket Médio"])
    for _, row in df_hora.iterrows():
        ws.append([row["HORA_STR"], float(row["TOTAL"]), int(row["COD_VENDA"]), float(row["TICKET_MEDIO"])])
    for col in ["B", "C", "D"]:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal="center")
    wb.save(output)

    st.download_button(
        label="📥 Baixar Excel",
        data=output.getvalue(),
        file_name="vendas_por_hora.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
#===========================================================================================================================================================

#Evolução de venda por dia da semana
#===========================================================================================================================================================
with st.container(border=True):
    st.markdown("<h4 style='color:#862E3A;'>📊 Evolução de Faturamento por Dia da Semana (Drilldown Mensal com Cores)</h4>", unsafe_allow_html=True)

    df_filt["MES_ANO"] = df_filt["DATA"].dt.to_period("M").astype(str)
    meses_disp = sorted(df_filt["MES_ANO"].unique())

    # Selecionar mês atual e anterior por padrão
    if len(meses_disp) >= 2:
        default_meses = [meses_disp[-2], meses_disp[-1]]
    else:
        default_meses = meses_disp

    meses_selecionados = st.multiselect("Selecionar Mês(es):", meses_disp, default=default_meses)

    df_mes = df_filt[df_filt["MES_ANO"].isin(meses_selecionados)].copy()
    df_mes["SEMANA"] = df_mes["DATA"].dt.isocalendar().week
    df_mes["ANO"] = df_mes["DATA"].dt.year
    dias_traduzidos = {
        "Sunday": "domingo", "Monday": "segunda-feira", "Tuesday": "terça-feira", "Wednesday": "quarta-feira",
        "Thursday": "quinta-feira", "Friday": "sexta-feira", "Saturday": "sábado"
    }
    df_mes["DIA_SEMANA"] = df_mes["DATA"].dt.day_name().map(dias_traduzidos)

    df_mes["INICIO_SEMANA"] = df_mes["DATA"] - pd.to_timedelta(df_mes["DATA"].dt.weekday, unit="d")
    df_mes["FIM_SEMANA"] = df_mes["INICIO_SEMANA"] + pd.Timedelta(days=6)
    df_mes["PERIODO"] = df_mes["INICIO_SEMANA"].dt.strftime('%d/%m') + " à " + df_mes["FIM_SEMANA"].dt.strftime('%d/%m')

    df_grouped = df_mes.groupby(["SEMANA", "PERIODO", "DIA_SEMANA"])["TOTAL"].sum().reset_index()
    df_pivot = df_grouped.pivot(index="DIA_SEMANA", columns="PERIODO", values="TOTAL").fillna(0)

    # Ordem começando por domingo
    ordem = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado", "domingo"]
    df_pivot = df_pivot.reindex(ordem)
    df_pivot = df_pivot[sorted(df_pivot.columns, key=lambda x: datetime.strptime(x.split(" à ")[0], "%d/%m"))]

    df_formatada = pd.DataFrame(index=df_pivot.index)
    colunas = df_pivot.columns.tolist()
    variacoes_pct = pd.DataFrame(index=df_pivot.index)

    for i, col in enumerate(colunas):
        col_fmt = []
        var_list = []
        for idx in df_pivot.index:
            valor = df_pivot.loc[idx, col]
            texto = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            variacao = None

            if i > 0:
                valor_ant = df_pivot.loc[idx, colunas[i - 1]]
                if valor_ant > 0:
                    variacao = (valor - valor_ant) / valor_ant
                    cor = "green" if variacao > 0 else "red"
                    texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
            col_fmt.append(texto)
            var_list.append(variacao)
        df_formatada[col] = col_fmt
        variacoes_pct[col] = var_list

    # === Totais semanais
    totais_semana = df_pivot.sum(axis=0)
    totais_variacoes = []
    totais_formatados = []

    for i, col in enumerate(colunas):
        valor = totais_semana[col]
        texto = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        variacao = None
        if i > 0:
            valor_ant = totais_semana[colunas[i - 1]]
            if valor_ant > 0:
                variacao = (valor - valor_ant) / valor_ant
                cor = "green" if variacao > 0 else "red"
                texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
        totais_formatados.append(texto)
        totais_variacoes.append(variacao)

    # === Tabela HTML com linha extra de totais
    tabela_html = "<table style='border-collapse: collapse; width: 100%; text-align: center;'>"
    tabela_html += "<thead><tr><th style='padding: 6px; border: 1px solid #555;'>DIA_SEMANA</th>"

    for col in colunas:
        tabela_html += f"<th style='padding: 6px; border: 1px solid #555;'>{col}</th>"
    tabela_html += "</tr></thead><tbody>"

    for idx in df_formatada.index:
        tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold'>{idx}</td>"
        for col in colunas:
            celula = df_formatada.loc[idx, col]
            pct = variacoes_pct.loc[idx, col]
            if pct is None or pd.isna(pct):
                fundo = "#f0f0f0"
            elif pct >= 0:
                fundo = "#CCFFCC"
            else:
                fundo = "#FFCCCC"
            tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111;'>{celula}</td>"
        tabela_html += "</tr>"

    # Linha de totais
    tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold; background-color: #ddd; color: #111;'>TOTAL</td>"
    for i, col in enumerate(colunas):
        fundo = "#f0f0f0"
        if i > 0 and totais_variacoes[i] is not None:
            fundo = "#CCFFCC" if totais_variacoes[i] >= 0 else "#FFCCCC"
        tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111; font-weight: bold;'>{totais_formatados[i]}</td>"


    st.markdown(tabela_html, unsafe_allow_html=True)

    # === Exportação para Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"
    
    # Cabeçalhos
    ws.append(["DIA_SEMANA"] + colunas)
    
    # Dados da tabela
    for idx in df_pivot.index:
        linha = [idx]
        for col in colunas:
            val = df_pivot.loc[idx, col]
            linha.append(round(val, 2))
        ws.append(linha)
    
    # Linha de total ao final
    linha_total = ["TOTAL"]
    for col in colunas:
        val = totais_semana[col]
        linha_total.append(round(val, 2))
    ws.append(linha_total)
    
    # Estilização no Excel
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row):
        for cell in row:
            pct_row = cell.row - 2
            pct_col = cell.column - 2
            try:
                is_total = cell.row == ws.max_row
    
                if is_total:
                    pct = totais_variacoes[pct_col] if pct_col < len(totais_variacoes) else None
                else:
                    pct = variacoes_pct.iloc[pct_row, pct_col]
    
                if pct is not None:
                    if pct >= 0:
                        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.fill = fill
    
                cell.font = Font(color="000000")  # cor preta para todas as células
                cell.alignment = Alignment(horizontal="center")
    
            except Exception:
                continue
                    
    # Download
    wb.save(output)
    st.download_button(
        label="📥 Baixar Excel",
        data=output.getvalue(),
        file_name="comparativo_dia_da_semana.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
#===========================================================================================================================================================


# Evolução da quantidade de vendas por dia da semana
#===========================================================================================================================================================
with st.container(border=True):
    st.markdown("<h4 style='color:#862E3A;'>🧾 Evolução de Quantidade de Vendas por Dia da Semana (Drilldown Mensal com Cores)</h4>", unsafe_allow_html=True)

    df_filt["MES_ANO"] = df_filt["DATA"].dt.to_period("M").astype(str)
    meses_disp = sorted(df_filt["MES_ANO"].unique())

    if len(meses_disp) >= 2:
        default_meses = [meses_disp[-2], meses_disp[-1]]
    else:
        default_meses = meses_disp

    meses_qtd = st.multiselect("Selecionar Mês(es):", meses_disp, default=default_meses, key="meses_qtd_venda")

    df_mes = df_filt[df_filt["MES_ANO"].isin(meses_qtd)].copy()
    df_mes["SEMANA"] = df_mes["DATA"].dt.isocalendar().week
    df_mes["ANO"] = df_mes["DATA"].dt.year
    dias_traduzidos = {
        "Sunday": "domingo", "Monday": "segunda-feira", "Tuesday": "terça-feira", "Wednesday": "quarta-feira",
        "Thursday": "quinta-feira", "Friday": "sexta-feira", "Saturday": "sábado"
    }
    df_mes["DIA_SEMANA"] = df_mes["DATA"].dt.day_name().map(dias_traduzidos)
    df_mes["INICIO_SEMANA"] = df_mes["DATA"] - pd.to_timedelta(df_mes["DATA"].dt.weekday, unit="d")
    df_mes["FIM_SEMANA"] = df_mes["INICIO_SEMANA"] + pd.Timedelta(days=6)
    df_mes["PERIODO"] = df_mes["INICIO_SEMANA"].dt.strftime('%d/%m') + " à " + df_mes["FIM_SEMANA"].dt.strftime('%d/%m')

    # Agrupamento
    df_grouped = df_mes.groupby(["SEMANA", "PERIODO", "DIA_SEMANA"])["COD_VENDA"].nunique().reset_index()
    df_grouped.rename(columns={"COD_VENDA": "QTDE"}, inplace=True)
    df_pivot = df_grouped.pivot(index="DIA_SEMANA", columns="PERIODO", values="QTDE").fillna(0)

    ordem = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado", "domingo"]
    df_pivot = df_pivot.reindex(ordem)
    df_pivot = df_pivot[sorted(df_pivot.columns, key=lambda x: datetime.strptime(x.split(" à ")[0], "%d/%m"))]

    df_formatada = pd.DataFrame(index=df_pivot.index)
    colunas = df_pivot.columns.tolist()
    variacoes_pct = pd.DataFrame(index=df_pivot.index)

    for i, col in enumerate(colunas):
        col_fmt = []
        var_list = []
        for idx in df_pivot.index:
            valor = df_pivot.loc[idx, col]
            texto = f"{int(valor):,}".replace(",", ".")
            variacao = None

            if i > 0:
                valor_ant = df_pivot.loc[idx, colunas[i - 1]]
                if valor_ant > 0:
                    variacao = (valor - valor_ant) / valor_ant
                    cor = "green" if variacao > 0 else "red"
                    texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
            col_fmt.append(texto)
            var_list.append(variacao)
        df_formatada[col] = col_fmt
        variacoes_pct[col] = var_list

    totais_semana = df_pivot.sum(axis=0)
    totais_variacoes = []
    totais_formatados = []

    for i, col in enumerate(colunas):
        valor = totais_semana[col]
        texto = f"{int(valor):,}".replace(",", ".")
        variacao = None
        if i > 0:
            valor_ant = totais_semana[colunas[i - 1]]
            if valor_ant > 0:
                variacao = (valor - valor_ant) / valor_ant
                cor = "green" if variacao > 0 else "red"
                texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
        totais_formatados.append(texto)
        totais_variacoes.append(variacao)

    # === Tabela HTML
    tabela_html = "<table style='border-collapse: collapse; width: 100%; text-align: center;'>"
    tabela_html += "<thead><tr><th style='padding: 6px; border: 1px solid #555;'>DIA_SEMANA</th>"

    for col in colunas:
        tabela_html += f"<th style='padding: 6px; border: 1px solid #555;'>{col}</th>"
    tabela_html += "</tr></thead><tbody>"

    for idx in df_formatada.index:
        tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold'>{idx}</td>"
        for col in colunas:
            celula = df_formatada.loc[idx, col]
            pct = variacoes_pct.loc[idx, col]
            if pct is None or pd.isna(pct):
                fundo = "#f0f0f0"
            elif pct >= 0:
                fundo = "#CCFFCC"
            else:
                fundo = "#FFCCCC"
            tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111;'>{celula}</td>"
        tabela_html += "</tr>"

    # Linha de totais
    tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold; background-color: #ddd; color: #111;'>TOTAL</td>"
    for i, col in enumerate(colunas):
        fundo = "#f0f0f0"
        if i > 0 and totais_variacoes[i] is not None:
            fundo = "#CCFFCC" if totais_variacoes[i] >= 0 else "#FFCCCC"
        tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111; font-weight: bold;'>{totais_formatados[i]}</td>"
    tabela_html += "</tr></tbody></table>"

    st.markdown(tabela_html, unsafe_allow_html=True)

    # === Exportação para Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Qtd_Vendas"

    ws.append(["DIA_SEMANA"] + colunas)
    for idx in df_pivot.index:
        linha = [idx]
        for col in colunas:
            val = df_pivot.loc[idx, col]
            linha.append(int(val))
        ws.append(linha)

    linha_total = ["TOTAL"]
    for col in colunas:
        val = totais_semana[col]
        linha_total.append(int(val))
    ws.append(linha_total)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row):
        for cell in row:
            pct_row = cell.row - 2
            pct_col = cell.column - 2
            try:
                is_total = cell.row == ws.max_row

                if is_total:
                    pct = totais_variacoes[pct_col] if pct_col < len(totais_variacoes) else None
                else:
                    pct = variacoes_pct.iloc[pct_row, pct_col]

                if pct is not None:
                    if pct >= 0:
                        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.fill = fill

                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center")

            except Exception:
                continue

    wb.save(output)
    st.download_button(
        label="📥 Baixar Excel (Qtd de Vendas)",
        data=output.getvalue(),
        file_name="comparativo_qtd_dia_da_semana.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
#===========================================================================================================================================================

# Evolução do Ticket Médio por Dia da Semana
#===========================================================================================================================================================
with st.container(border=True):
    st.markdown("<h4 style='color:#862E3A;'>💳 Evolução do Ticket Médio por Dia da Semana (Drilldown Mensal com Cores)</h4>", unsafe_allow_html=True)

    df_filt["MES_ANO"] = df_filt["DATA"].dt.to_period("M").astype(str)
    meses_disp = sorted(df_filt["MES_ANO"].unique())

    if len(meses_disp) >= 2:
        default_meses = [meses_disp[-2], meses_disp[-1]]
    else:
        default_meses = meses_disp

    meses_ticket = st.multiselect("Selecionar Mês(es):", meses_disp, default=default_meses, key="meses_ticket_medio")

    df_mes = df_filt[df_filt["MES_ANO"].isin(meses_ticket)].copy()
    df_mes["SEMANA"] = df_mes["DATA"].dt.isocalendar().week
    df_mes["ANO"] = df_mes["DATA"].dt.year
    dias_traduzidos = {
        "Sunday": "domingo", "Monday": "segunda-feira", "Tuesday": "terça-feira", "Wednesday": "quarta-feira",
        "Thursday": "quinta-feira", "Friday": "sexta-feira", "Saturday": "sábado"
    }
    df_mes["DIA_SEMANA"] = df_mes["DATA"].dt.day_name().map(dias_traduzidos)
    df_mes["INICIO_SEMANA"] = df_mes["DATA"] - pd.to_timedelta(df_mes["DATA"].dt.weekday, unit="d")
    df_mes["FIM_SEMANA"] = df_mes["INICIO_SEMANA"] + pd.Timedelta(days=6)
    df_mes["PERIODO"] = df_mes["INICIO_SEMANA"].dt.strftime('%d/%m') + " à " + df_mes["FIM_SEMANA"].dt.strftime('%d/%m')

    df_grouped = df_mes.groupby(["SEMANA", "PERIODO", "DIA_SEMANA"]).agg({"TOTAL": "sum", "COD_VENDA": "nunique"}).reset_index()
    df_grouped["TICKET"] = df_grouped["TOTAL"] / df_grouped["COD_VENDA"]
    df_pivot = df_grouped.pivot(index="DIA_SEMANA", columns="PERIODO", values="TICKET").fillna(0)

    ordem = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira", "sexta-feira", "sábado", "domingo"]
    df_pivot = df_pivot.reindex(ordem)
    df_pivot = df_pivot[sorted(df_pivot.columns, key=lambda x: datetime.strptime(x.split(" à ")[0], "%d/%m"))]

    df_formatada = pd.DataFrame(index=df_pivot.index)
    colunas = df_pivot.columns.tolist()
    variacoes_pct = pd.DataFrame(index=df_pivot.index)

    for i, col in enumerate(colunas):
        col_fmt = []
        var_list = []
        for idx in df_pivot.index:
            valor = df_pivot.loc[idx, col]
            texto = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            variacao = None

            if i > 0:
                valor_ant = df_pivot.loc[idx, colunas[i - 1]]
                if valor_ant > 0:
                    variacao = (valor - valor_ant) / valor_ant
                    cor = "green" if variacao > 0 else "red"
                    texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
            col_fmt.append(texto)
            var_list.append(variacao)
        df_formatada[col] = col_fmt
        variacoes_pct[col] = var_list

    totais_semana = df_pivot.mean(axis=0)
    totais_variacoes = []
    totais_formatados = []

    for i, col in enumerate(colunas):
        valor = totais_semana[col]
        texto = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        variacao = None
        if i > 0:
            valor_ant = totais_semana[colunas[i - 1]]
            if valor_ant > 0:
                variacao = (valor - valor_ant) / valor_ant
                cor = "green" if variacao > 0 else "red"
                texto += f"<br><span style='color:{cor}; font-size: 12px'>{variacao:+.2%}</span>"
        totais_formatados.append(texto)
        totais_variacoes.append(variacao)

    tabela_html = "<table style='border-collapse: collapse; width: 100%; text-align: center;'>"
    tabela_html += "<thead><tr><th style='padding: 6px; border: 1px solid #555;'>DIA_SEMANA</th>"

    for col in colunas:
        tabela_html += f"<th style='padding: 6px; border: 1px solid #555;'>{col}</th>"
    tabela_html += "</tr></thead><tbody>"

    for idx in df_formatada.index:
        tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold'>{idx}</td>"
        for col in colunas:
            celula = df_formatada.loc[idx, col]
            pct = variacoes_pct.loc[idx, col]
            if pct is None or pd.isna(pct):
                fundo = "#f0f0f0"
            elif pct >= 0:
                fundo = "#CCFFCC"
            else:
                fundo = "#FFCCCC"
            tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111;'>{celula}</td>"
        tabela_html += "</tr>"

    # Linha de totais
    tabela_html += f"<tr><td style='padding: 6px; border: 1px solid #555; font-weight: bold; background-color: #ddd; color: #111;'>TOTAL</td>"
    for i, col in enumerate(colunas):
        fundo = "#f0f0f0"
        if i > 0 and totais_variacoes[i] is not None:
            fundo = "#CCFFCC" if totais_variacoes[i] >= 0 else "#FFCCCC"
        tabela_html += f"<td style='padding: 6px; border: 1px solid #555; background-color: {fundo}; color: #111; font-weight: bold;'>{totais_formatados[i]}</td>"
    tabela_html += "</tr></tbody></table>"

    st.markdown(tabela_html, unsafe_allow_html=True)

    # Export Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ticket_Medio"

    ws.append(["DIA_SEMANA"] + colunas)
    for idx in df_pivot.index:
        linha = [idx]
        for col in colunas:
            val = df_pivot.loc[idx, col]
            linha.append(round(val, 2))
        ws.append(linha)

    linha_total = ["TOTAL"]
    for col in colunas:
        val = totais_semana[col]
        linha_total.append(round(val, 2))
    ws.append(linha_total)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row):
        for cell in row:
            pct_row = cell.row - 2
            pct_col = cell.column - 2
            try:
                is_total = cell.row == ws.max_row

                if is_total:
                    pct = totais_variacoes[pct_col] if pct_col < len(totais_variacoes) else None
                else:
                    pct = variacoes_pct.iloc[pct_row, pct_col]

                if pct is not None:
                    if pct >= 0:
                        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.fill = fill

                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center")

            except Exception:
                continue

    wb.save(output)
    st.download_button(
        label="📥 Baixar Excel (Ticket Médio)",
        data=output.getvalue(),
        file_name="comparativo_ticket_medio.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
#===========================================================================================================================================================



# =======================
# 📋 TABELA DETALHADA
# =======================
with st.expander("📋 Ver dados detalhados"):
    st.markdown("### 📄 Dados Filtrados por UN e Período Selecionado")
    st.dataframe(df_filt, use_container_width=True)
